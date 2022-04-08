# Voor de GUI
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
# Tijd op te halen
from datetime import datetime
# Excel gezeik
import openpyxl
from openpyxl import Workbook
import pathlib
# Voor de wait
import time

# Template voor data.txt:
# Tijd, Barcode, Naam, Goed/afgekeurd, Comment, Productiestap
# newline hoef niet toegevoegd te worden dat doet dit programma zelf.

# dit programma naar exe: https://www.youtube.com/watch?v=QWqxRchawZY

root = Tk()
root.title('Barcodescanner Protoworkz')
# Protoworkz logo | werkt op een of andere vage reden niet
#root.iconbitmap('protoworkz.ico')
# past automatisch het venster aan | bleekt toch kut te zijn
#root.geometry("{0}x{1}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))

FASTSLEEP = 0.2
SLEEP = 2
LONGSLEEP = 10

# optie om data ook naar een txt bestand over tezetten (voor het geval je toch wil controleren maar er is geen ecxel beschikbaar, hij export alles wel nog naar excel ongeacht deze optie)
txtfileexport = True

# kut excel gezeik
excelbestand = pathlib.Path('data.xlsx')
# checkt of het bestand al bestand zo niet maakt hij een bestand met de benodigde columtjes
if excelbestand.exists ():
    pass
else:
    excelbestand=Workbook()
    sheet=excelbestand.active
    sheet["A1"]='Tijd'
    sheet["B1"]='Barcode'
    sheet["C1"]='Naam'
    sheet["D1"]='Goed/Afgekeurd'
    sheet["E1"]='Comment'
    sheet["F1"]='Productiestap'

    excelbestand.save('data.xlsx')


def destroyAllButtons():
    myButton2.destroy()
    myButtonBarcode.destroy()
    myButtonDestroy.destroy()
    myButtonZoeken.destroy()

def destroyAllLabels():
    myLabelKiesNaam.destroy()

def onClickBarcode():
    # dropdown menu van: https://stackoverflow.com/questions/45441885/how-can-i-create-a-dropdown-menu-from-a-list-in-tkinter
    time.sleep(FASTSLEEP)
    destroyAllButtons()
    destroyAllLabels()
    dropNaam.destroy()
    # myLabelBarcodescannen = Label(root, text='Selecteer het vak hieronder en scan vervolgens de barcode')
    # myLabelBarcodescannen.grid(row=0, column=5)
    # myLabelKiesding = Label(root, text='Selecteer productiestap')
    # myLabelKiesding.grid(row=0, column=6)
    e = Entry(root)
    # e.insert(0, 'Geef hier je feedback indien afgekeurd')
    e.grid(row=1, column=8, ipady=20, ipadx=50, sticky=N)

    options = [
    # '-',
    'Reflow',
    'THT',
    'Test en Prog',
    'Laatste controle'
    ]

    clicked = StringVar()
    clicked.set(options[0])

    drop = OptionMenu(root, clicked, *options)
    drop.grid(row=0, column=6)
    drop.config(width=40, height=10)

    myCombo = ttk.Combobox(root, value=options)
    myCombo.current(0)
    myCombo.bind('Comboboxselected')

    def extract_data():
        # dat file gezeik heb ik hier vandaan: https://www.w3schools.com/python/python_file_handling.asp
        barcodenummer = text_box.get('1.0', 'end')
        now = datetime.now()
        current_time = now.strftime("%d/%m/%Y | %H:%M:%S")
        file = open('data.txt', 'a')
        if barcodenummer > '10':
            if var1.get() == 1 and not var2.get() == 1 and not clicked.get() == '-':
                excelbestand=openpyxl.load_workbook('data.xlsx')
                sheet=excelbestand.active
                # pleurt alles in goede volgorde in excel bestand
                sheet.cell(column=1,row=sheet.max_row+1,value=current_time)
                sheet.cell(column=2,row=sheet.max_row,value=barcodenummer)
                sheet.cell(column=3,row=sheet.max_row,value=clickedNaam.get())
                sheet.cell(column=4,row=sheet.max_row,value='Goedgekeurd')
                sheet.cell(column=5,row=sheet.max_row,value='')
                sheet.cell(column=6,row=sheet.max_row,value=clicked.get())
                if txtfileexport:
                    file.write('\n')
                    file.write(current_time)
                    file.write(',')
                    file.write(barcodenummer.strip())
                    file.write(',')
                    file.write(clickedNaam.get())
                    file.write(',')
                    file.write("Goedgekeurd")
                    file.write(',')
                    file.write(' ')
                    file.write(',')
                    file.write(clicked.get())
                print(current_time, barcodenummer.strip(), 'toegevoegd aan', file.name, 'toegevoegd door:', clickedNaam.get(), 'Product is Goedgekeurd', ', Stap:', clicked.get())
                myLabelSuccesvol = Label(root, text='Barcode succesvol in bestand gezet.')
                myLabelSuccesvol.grid(row=1, column=11, sticky=N)
                #destroyed na 3000ms de label
                root.after(3000, lambda: myLabelSuccesvol.destroy())
                # delete de tekst vlakken als het is toegevoegd
                e.delete(0, 'end')
                text_box.delete(1.0, "end")
                # die save hieronder is essentieel
                excelbestand.save('data.xlsx')
            elif var2.get() == 1 and not var1.get() == 1:
                if e.get() > '5' and not e.get() == 'Geef hier je feedback indien afgekeurd' and not clicked.get() == '-':
                    excelbestand=openpyxl.load_workbook('data.xlsx')
                    sheet=excelbestand.active
                    # pleurt alles in goede volgorde in excel bestand
                    sheet.cell(column=1,row=sheet.max_row+1,value=current_time)
                    sheet.cell(column=2,row=sheet.max_row,value=barcodenummer)
                    sheet.cell(column=3,row=sheet.max_row,value=clickedNaam.get())
                    sheet.cell(column=4,row=sheet.max_row,value='Afgekeurd')
                    sheet.cell(column=5,row=sheet.max_row,value=e.get())
                    sheet.cell(column=6,row=sheet.max_row,value=clicked.get())
                    if txtfileexport:
                        file.write('\n')
                        file.write(current_time)
                        file.write(',')
                        file.write(barcodenummer.strip())
                        file.write(',')
                        file.write(clickedNaam.get())
                        file.write(',')
                        file.write("Afgekeurd")
                        file.write(',')
                        file.write(e.get())
                        file.write(',')
                        file.write(clicked.get())
                    # delete de tekst vlakken als het is toegevoegd
                    e.delete(0, 'end')
                    text_box.delete(1.0, "end")
                    # die save hieronder is essentieel
                    excelbestand.save('data.xlsx')
                    print(current_time, barcodenummer.strip(), 'toegevoegd aan', file.name, 'toegevoegd door:', clickedNaam.get(), 'Product is Afgekeurd met de reden:', e.get(), ', Stap: ', clicked.get())
                    myLabelSuccesvol = Label(root, text='Barcode succesvol in bestand gezet.')
                    myLabelSuccesvol.grid(row=1, column=11, sticky=N)
                    #destroyed na 3000ms de label
                    root.after(3000, lambda: myLabelSuccesvol.destroy())
                elif clicked.get() == '-':
                    print('Selecteer de productie stap!')
                    # myLabelProductiestap = Label(root, text='Selecteer de productie stap!')
                    # myLabelProductiestap.grid(row=7, column=11)
                    messagebox.showerror("Barcode scanner Protoworkz", "Selecteer de productie stap!")
                else:
                    print('Feedback is vereist! (Feedback mag geen \'-\' bevatten!)')
                    # myLabelFeedbackfoei = Label(root, text='Feedback is vereist!')
                    # myLabelFeedbackfoei.grid(row=7, column=11)
                    messagebox.showerror("Barcode scanner Protoworkz", "Feedback is vereist! (Feedback mag geen \'-\' bevatten!)")
            else:
                print('Er is iets mis gegaan tijdens het goed/afkeuren!')
                # myLabelGoedafkeurfoei = Label(root, text='Er is iets mis gegaan tijdens het goed/afkeuren!')
                # myLabelGoedafkeurfoei.grid(row=7, column=11)
                messagebox.showerror("Barcode scanner Protoworkz", "Er is iets mis gegaan tijdens het goed/afkeuren!")
                return
        else:
            # error boxen komen van: https://www.geeksforgeeks.org/python-tkinter-messagebox-widget/
            # de labels die zijn gecomment overlappen in elkaar en weet geen fix dr voor
            print('Ongeldige barcode.')
            # myLabelOngeldigeBarcode = Label(root, text='Ongeldige barcode (11 cijfers).')
            # myLabelOngeldigeBarcode.grid(row=7, column=11)
            messagebox.showerror("Barcode scanner Protoworkz", "Ongeldige barcode.")
            return

    # textbox van https://pythonguides.com/python-tkinter-text-box/
    message =''

    text_box = Text(
        root,
        height=13,
        width=80,
        wrap='word'
    )
    text_box.grid(row=0, column=5)
    text_box.insert('end', message)

    knopBarcode = Button(
        root,
        width=30,
        height=15,
        fg='white',
        bg='darkgrey',
        text='Barcode in bestand zetten',
        command=extract_data
    ).grid(row=0, column=11)

    # Label(root, text="Goed of afgekeurd:").grid(row=0, sticky=W, column=8)
    var1 = IntVar()
    Checkbutton(root, text="Goedgekeurd", fg='green', variable=var1).grid(row=0, column=8)
    var2 = IntVar()
    Checkbutton(root, text="Afgekeurd", fg='red', variable=var2).grid(row=0, sticky=S, column=8)

    # resize van codemy.com | Dynamically Resize Buttons When Resizing a Window - Python Tkinter GUI tutorial #145
    lijstMetKnoppen3 = [knopBarcode, text_box, e]

    row_number3 = 0

    for knop in lijstMetKnoppen3:
        Grid.rowconfigure(root, row_number3, weight=1)
        row_number3 += 1

def destroyWindow():
    time.sleep(FASTSLEEP)
    root.destroy()
    print('--Venster is gesloten.--')

def onClickLezen():
    time.sleep(SLEEP)
    file = open('data.txt', 'r')
    print(file.read())
    # time.sleep(SLEEP)
    destroyWindow()

def onClickZoeken():
    destroyAllButtons()
    destroyAllLabels()
    dropNaam.destroy()

    def dataZoeken():
        # deze functie heb ik hiervandaan: https://www.geeksforgeeks.org/python-program-to-print-lines-containing-given-string-in-file/
        zoeken = text_boxzoek.get('1.0', 'end')
        file_name = "data.txt"
        try:
    
            file_read = open(file_name, "r")
        
            text = zoeken.strip()
        
            lines = file_read.readlines()
        
            new_list = []
            idx = 0
        
            for line in lines:
                
                if text in line:
                    new_list.insert(idx, line)
                    idx += 1
        
            file_read.close()
        
            if len(new_list)==0:
                # myLabelNietgevonden = Label(root, text="\n\"" +text+ "\" is niet gevonden in \"" +file_name+ "\"!")
                # myLabelNietgevonden.grid(row=5, column=1)
                print("\n\"" +text+ "\" is niet gevonden in \"" +file_name+ "\"!")
                # root.after(10000, lambda: myLabelNietgevonden.destroy())
            else:
        
                lineLen = len(new_list)
                # myLabelGevonden = Label(root, text="\n**** Lines die \"" +text+ "\" bevatten ****\n")
                # myLabelGevonden.grid(row=7, column=11)
                print("\n**** Lines die \"" +text+ "\" bevatten ****\n")
                # root.after(10000, lambda: myLabelGevonden.destroy())
                for i in range(lineLen):
                    print(end=new_list[i])
                    # myLabelGevonden = Label(root, text=new_list[i])
                    # myLabelGevonden.grid(row=7, column=1)
                    # root.after(10000, lambda: myLabelGevonden.destroy())
        
        except:
            print("\nBestand bestaat niet!")

    # textbox van https://pythonguides.com/python-tkinter-text-box/
    message =''

    text_boxzoek = Text(
        root,
        height=13,
        width=80,
        wrap='word'
    )
    text_boxzoek.grid(row=1, column=5)
    text_boxzoek.insert('end', message)

    knopZoeken = Button(
        root,
        width=30,
        height=15,
        fg='black',
        bg='yellow',
        text='Barcode opzoeken',
        command=dataZoeken
    ).grid(row=2, column=11)

# resize van codemy.com | Dynamically Resize Buttons When Resizing a Window - Python Tkinter GUI tutorial #145
    lijstMetKnoppen2 = [knopZoeken]

    row_number2 = 0

    for knop in lijstMetKnoppen2:
        Grid.rowconfigure(root, row_number2, weight=1)
        row_number2 += 1


myLabelKiesNaam = Label(root, text='Selecteer uw naam')
myLabelKiesNaam.grid(row=0, column=9)


# dropdown menu van: https://stackoverflow.com/questions/45441885/how-can-i-create-a-dropdown-menu-from-a-list-in-tkinter
optionsNaam = [
'Robert',
'Max',
'Bastiaan'
]

clickedNaam = StringVar()
clickedNaam.set(optionsNaam[0])

dropNaam = OptionMenu(root, clickedNaam, *optionsNaam)
dropNaam.grid(row=1, column=9)
dropNaam.config(width=40, height=20)

myComboNaam = ttk.Combobox(root, value=optionsNaam)
myComboNaam.current(0)
myComboNaam.bind('Comboboxselected')

myButtonBarcode = Button(root, text='Barcode inlezen', width=20, height=10, command=onClickBarcode, fg='white', bg='red')
myButtonBarcode.grid(row=2, column=4)
myButton2 = Button(root, text='Bestand lezen', width=20, height=10, command=onClickLezen, fg='white', bg='grey')
myButton2.grid(row=2, column=6)
myButtonZoeken = Button(root, text='Barcode zoeken', width=20, height=10, command=onClickZoeken, fg='white', bg='grey')
myButtonZoeken.grid(row=2, column=8)
myButtonDestroy = Button(root, text='Venster sluiten', width=20, height=10, command=destroyWindow, fg='black')
myButtonDestroy.grid(row=2, column=10)


# resize van codemy.com | Dynamically Resize Buttons When Resizing a Window - Python Tkinter GUI tutorial #145
lijstMetKnoppenBegin = [myButton2, myButtonBarcode, myButtonDestroy, myButtonZoeken]

row_number = 0

for knop in lijstMetKnoppenBegin:
    Grid.rowconfigure(root, row_number, weight=1)
    row_number += 1
# myButton.pack()
# myLabelNaam = Label(root, text=naam)
# myLabelNaam.grid(row=0, column=4)
# myLabel = Label(root, text='Wat wil je gaan doen?')
# myLabel.grid(row=0, column=7)

# keuze = input('Wat wil je in het document hebben: ')

# if keuze == 'toevoegen' or keuze == 't':
#     antwoord = str(input('Wat wil je toegevoegd hebben: '))
#     file = open('test.txt', 'a')
#     file.write('\n')
#     file.write(antwoord)
#     print(antwoord, 'toegevoegd aan', file)
#     file.close()
# else:
#     print('Kan opdracht niet uitvoeren')

root.mainloop()
